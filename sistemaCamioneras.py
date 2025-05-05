# Importación de Librerias
import serial
import sys
import time
import threading
import re
import ast
from datetime import datetime
from PyQt5.QtWidgets import QMainWindow, QLabel, QTableWidgetItem, QWidget, QHBoxLayout, QApplication, QDialog, QAction,QFileDialog, QMessageBox
from PyQt5.QtCore import QThread, pyqtSignal, Qt, QDate, QObject, QEvent
from PyQt5 import QtCore, QtWidgets, QtGui, uic
from PyQt5.QtGui import QPixmap, QMovie, QColor, QFont
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from reportlab.lib.styles import getSampleStyleSheet
import win32print
import win32api
import win32ui
import fitz
import subprocess
import os
import openpyxl
from openpyxl.styles import Alignment
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

# Importación de los Layout
from View.ui_Principal import Ui_viewInicioApp # La clase Ui_MainWindow del archivo ui_Principal.py 
from View.Ui_modal_reporte import Ui_modal_reporte

# Importación de Base de Datos
from Base_de_Datos.database_conexion import Conectar # El archivo database_conexion.py

# Puertos COM
COM1 = ""
COMAR = ""

# Corte Cadena
cadenaPrincipio = 0
cadenaFinal = 0

# Variables Bandera
capturarPesoBruto = False
capturarPesoTara = False

# Variables de rutas de imagenes para alerta
correcto = "Resources/correcto.png"
error = "Resources/error.png"

# Variable nombre de la Impresora
nombreImpresora = ""
nombreEmpresa = ""
ubicacionEmpresa = ""
rucEmpresa = ""

class PDFViewer(QtWidgets.QWidget):
    def __init__(self, file_path):
        super().__init__()
        self.file_path = file_path
        self.init_ui()
        
    def init_ui(self):
        self.setWindowTitle('VISTA PREVIA DEL DOCUMENTO')
        self.setWindowIcon(QtGui.QIcon("Resources/icono.png"))
        self.resize(673, 600)
        self.center_window()

        layout = QtWidgets.QVBoxLayout(self)
        
        self.scroll_area = QtWidgets.QScrollArea(self)
        self.scroll_area.setWidgetResizable(True)
        
        self.pdf_widget = QtWidgets.QWidget()
        self.scroll_area.setWidget(self.pdf_widget)
        
        self.vbox = QtWidgets.QVBoxLayout(self.pdf_widget)
        
        layout.addWidget(self.scroll_area)
        
        self.load_pdf()
        
        self.print_button = QtWidgets.QPushButton("IMPRIMIR")
        self.print_button.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))

        # Crear y aplicar el QFont
        font = QtGui.QFont()
        font.setFamily("Montserrat")
        font.setPointSize(12)
        font.setBold(True)  # Para asegurarnos de que la fuente sea negrita

        self.print_button.setFont(font)

        # Aplicar el estilo CSS
        self.print_button.setStyleSheet("""
            QPushButton {
                background-color: #FFCC21;
                color: rgb(0, 0, 0);
                border: none;
                border-radius: 5px;
                border-left: 2px solid #C79D0F;
                border-right: 2px solid #C79D0F;
                border-bottom: 4px solid #C79D0F;
                padding: 5px;
            }
            QPushButton:pressed {
                background-color: #D5A80F;
            }
        """)
        self.print_button.clicked.connect(self.print_pdf)
        layout.addWidget(self.print_button)
    
    def center_window(self):
        qr = self.frameGeometry()
        cp = QtWidgets.QDesktopWidget().availableGeometry().center()
        qr.moveCenter(cp)
        self.move(qr.topLeft())
    
    def load_pdf(self):
        doc = fitz.open(self.file_path)
        for page_num in range(len(doc)):
            page = doc.load_page(page_num)
            pix = page.get_pixmap()
            img = QtGui.QImage(pix.samples, pix.width, pix.height, pix.stride, QtGui.QImage.Format_RGB888)
            img = img.rgbSwapped()
            lbl = QtWidgets.QLabel(self)
            lbl.setPixmap(QtGui.QPixmap.fromImage(img))
            self.vbox.addWidget(lbl)
    
    # def print_pdf(self):
    #     printer_name = win32print.GetDefaultPrinter()
    #     print(f"Imprimiendo en: {printer_name}")

    #     try:
    #         hprinter = win32print.OpenPrinter(printer_name)
    #         pdc = win32ui.CreateDC()
    #         pdc.CreatePrinterDC(printer_name)
    #         pdc.StartDoc("Impresion PDF Directa")
    #         pdc.StartPage()

            
    #         try:
    #             print("Enviando PDF a la impresora...")
    #             os.startfile(self.file_path, "print")
    #             print("PDF enviado correctamente a la impresora.")
    #         except Exception as e:
    #             print(f"Error al enviar el archivo PDF a la impresora: {e}")
    #         finally:
    #             pdc.EndPage()
    #             pdc.EndDoc()

    #     except Exception as e:
    #         print(f"Error al intentar imprimir: {e}")
    #     finally:
    #         win32print.ClosePrinter(hprinter)

    def print_pdf(self):
        try:
            printer_name = nombreImpresora
            printer_handle = win32print.OpenPrinter(printer_name)
            print("nombre de la imopresora ",printer_name )
            win32api.ShellExecute(
                0,
                "print",
                self.file_path,
                f'/d:"{printer_name}"',
                ".",
                0
            )
            win32print.ClosePrinter(printer_handle)
        except Exception as e:
            print("Error al imprimir el PDF:", e)

"""Creamos hilo para la ejecución en segundo plano del Indicador , de esta forma
evitamos que la aplicación se detenga por la lectura constante """

class WorkerThread(QThread):
    update_peso = pyqtSignal(str)
    update_estado = pyqtSignal(str)
    update_baliza = pyqtSignal(str)
    
    def run(self):
        COMINDICADOR1 = "COM" + COM1
        serialIndicador = None

        while True:
            try:
                if serialIndicador is None or not serialIndicador.is_open:
                    serialIndicador = serial.Serial(COMINDICADOR1, baudrate=9600, timeout=1, bytesize=serial.EIGHTBITS, parity=serial.PARITY_NONE, stopbits=serial.STOPBITS_ONE)
                
                import re  # Asegúrate de importar esto al inicio del archivo
                
                result = serialIndicador.readline().decode('utf-8', errors='ignore').strip()
                
                # Buscar una cadena que termine en "kg"
                #match = re.search(r'(wn\d+\.\d{2}kg)', result)
                #match = re.search(r'(wn\d+(?:\.\d{1,2})?kg)', result)
                match = re.search(r'(\d+)\s*KG', result.upper())

                
                if match:
                    numero = str(int(match.group(1)))
                    #clean_result = match.group(1)  # Extrae el texto limpio
                    self.update_peso.emit(numero)
                    self.update_baliza.emit(numero)
                    self.update_estado.emit("1")
                else:
                    self.update_peso.emit("-----")
                    self.update_baliza.emit("0.00")
                    self.update_estado.emit("0")
                    
            except Exception as e:
                # print("WT IN: " + str(e))
                time.sleep(1)
                # Cerrar la conexión serial si hay una excepción
                if serialIndicador is not None and serialIndicador.is_open:
                    serialIndicador.close()
    
    def stop(self):
        print("Thread Stopped")
        self.terminate()

""" Creamos hilo para la ejecución en segundo plano de la Fecha y Hora, de esta forma
evitamos que la aplicación se detenga por la lectura constante  """

class WorkerThreadFechaHora(QThread):
    
    update_fecha_hora = pyqtSignal(str)
    update_peso = pyqtSignal(str)
    update_baliza = pyqtSignal(str)
    update_estado = pyqtSignal(str)
    
    def run(self):
        COMINDICADOR1 = "COM" + COM1
        serialIndicador = None

        while True:
            hora_actual = datetime.now().time()
            hora = int(hora_actual.strftime("%H"))
            minutos = hora_actual.strftime("%M")
            segundos = hora_actual.strftime("%S")
            periodo = "AM" if hora < 12 else "PM"
            hora = hora if hora <= 12 else hora - 12
            hora_formateada = "{:02d} : {:02d} : {:02d} {}".format(hora, int(minutos), int(segundos), periodo)

            fecha_actual = datetime.now().date()
            año = fecha_actual.year
            mes = fecha_actual.month
            dia = fecha_actual.day
            dia_semana = fecha_actual.weekday()
            dia_semana = ["Lunes", 
                        "Martes", 
                        "Miércoles", 
                        "Jueves", 
                        "Viernes", 
                        "Sábado", 
                        "Domingo"][int(dia_semana)]
            meses = {
                1: "Enero",
                2: "Febrero",
                3: "Marzo",
                4: "Abril",
                5: "Mayo",
                6: "Junio",
                7: "Julio",
                8: "Agosto",
                9: "Septiembre",
                10: "Octubre",
                11: "Noviembre",
                12: "Diciembre"
            }
            fecha_formateada = "{} {} de {} del {}".format(dia_semana, dia, meses[mes], año)

            self.update_fecha_hora.emit(fecha_formateada + " - " + hora_formateada)
            time.sleep(1)
            try:
                # Intentar conectar si la conexión serial no está abierta
                if serialIndicador is None or not serialIndicador.is_open:
                    serialIndicador = serial.Serial(
                        port=COMINDICADOR1,
                        baudrate=9600,
                        timeout=1,
                        bytesize=serial.EIGHTBITS,
                        parity=serial.PARITY_NONE,
                        stopbits=serial.STOPBITS_ONE
                    )

                buffer = ""  # Acumula los datos recibidos

                while serialIndicador.is_open:
                    if serialIndicador.in_waiting > 0:
                        data = serialIndicador.read(serialIndicador.in_waiting)
                        try:
                            # Agregar los datos decodificados al buffer
                            buffer += data.decode('utf-8', errors='ignore')

                            # Procesar los datos completos dentro del buffer
                            while " " in buffer:
                                segment, buffer = buffer.split(" ", 1)

                                if segment.strip():
                                    segment = segment[:-10] if len(segment) > 10 else ""
                                    self.update_peso.emit(segment)
                                    self.update_baliza.emit(segment)
                                    self.update_estado.emit("1")
                        except UnicodeDecodeError:
                            print(f"Error al decodificar: {data}")

            except serial.SerialException as e:
                # print(f"Error en la conexión serial: {e}")
                time.sleep(1)
            except Exception as e:
                print(f"Error inesperado: {e}")
                time.sleep(1)
            finally:
                # Cerrar la conexión serial si ocurre una excepción
                if serialIndicador is not None and serialIndicador.is_open:
                    serialIndicador.close()

                # Emitir estado de desconexión
                self.update_peso.emit("-----")
                self.update_baliza.emit("0.00")
                self.update_estado.emit("0")

    def stop(self):
        print("Thread Stopped")
        self.terminate()
        
# ===============================
# Creación de la Clase Principal
# ===============================

class Inicio(QMainWindow):
    
    def __init__(self):
        super(Inicio, self).__init__()
        self.ui = Ui_viewInicioApp()
        self.ui.setupUi(self)
        
        self.modal_reporte = Ui_modal_reporte()
        
        self.setWindowIcon(QtGui.QIcon("Resources/icono.png"))
        self.setWindowTitle('Sistema de Pesaje || Balinsa')
        self.setWindowFlag(QtCore.Qt.FramelessWindowHint)
        
        self.workerFechaHora = WorkerThreadFechaHora() # Hilo de Fecha y Hora
        self.workerFechaHora.start() # Iniciamos el hilo
        self.workerFechaHora.update_fecha_hora.connect(self.mostrar_fecha_hora) # Llamamos a la función mostrar_fecha_hora

        self.conexion = Conectar()
        self.fn_asigna_recorte_cadena()
        self.fn_asigna_puertos_default()
        self.fn_asigna_nombreImpresora()
        
        self.ui.lblmimizar.setPixmap(QPixmap("Resources/minimizar.png"))
        self.ui.lblcerrar.setPixmap(QPixmap("Resources/cerrar.png"))
        # self.ui.imgLogoBalinsa.setPixmap(QPixmap("Resources/logoBalinsa.png"))
        self.ui.imgTriangulo.setPixmap(QPixmap("Resources/imgTriangulo.png"))
        self.ui.imgRectangulo.setPixmap(QPixmap("Resources/imgRectangulo.png"))
        self.ui.imgCancelar.setPixmap(QPixmap("Resources/cancelar_bn.png"))
        self.ui.imgCamion.setPixmap(QPixmap("Resources/camion_bn.png"))
        self.ui.imgGuardar.setPixmap(QPixmap("Resources/guardar_bn.png"))
        self.ui.imgImprimir.setPixmap(QPixmap("Resources/impresora_bn.png"))
        self.ui.imgBuscarTara.setPixmap(QPixmap("Resources/buscar.png"))
        self.ui.imgLogoEmpresa.setPixmap(QPixmap("Resources/imgLogoEmpresa.png"))
        self.ui.imgRefresh.setPixmap(QPixmap("Resources/refresh.png"))
        self.ui.imgEditBruto.setPixmap(QPixmap("Resources/edit.png"))
        self.ui.imgEditTara.setPixmap(QPixmap("Resources/edit.png"))
        self.ui.imgCalendario.setPixmap(QPixmap("Resources/calendario.png"))
        self.ui.imgRefresh.setHidden(True)
        self.ui.btnNuevoTicket.setHidden(True)
        self.ui.lblNuevoTicket.setHidden(True)
        
        self.ui.lblEstadoIndicador.setText("FUERA DE LINEA")
        self.ui.lblEstadoIndicador.setStyleSheet("background-color: rgb(234, 29, 49);color: rgb(255, 255, 255);border-radius: 5px;")
        self.ui.lblPesoIndicador.setText("-----")
        self.ui.lblGuardarActualizar.setText("Guardar")
        
        self.ui.calendarWidget.setHidden(True)
        self.ui.btnCalendario.clicked.connect(self.fn_abrir_calendario)
        self.ui.calendarWidget.clicked.connect(self.update_date)
        fechaActual = datetime.now().date()
        qdate = QDate(fechaActual.year, fechaActual.month, fechaActual.day)
        self.ui.dateEditConsultar.setDate(qdate)
        
        self.worker = WorkerThread() # Hilo Balanza
        self.worker.start()
        self.worker.update_peso.connect(self.evt_actualizar_peso)
        self.worker.update_estado.connect(self.evt_actualizar_estado)
        
        self.tablaDePesos = self.ui.tblDetallePesadas
        self.tablaDePesos.setColumnWidth(0, 111)
        self.tablaDePesos.setColumnWidth(1, 111)
        self.tablaDePesos.setColumnWidth(2, 111)
        self.tablaDePesos.setColumnWidth(3, 111)
        self.tablaDePesos.setColumnWidth(4, 111)
        self.tablaDePesos.setColumnWidth(5, 111)
        self.tablaDePesos.setColumnHidden(6, True)
        self.tablaDePesos.setAlternatingRowColors(True)
        self.tablaDePesos.clearContents()
        self.tablaDePesos.setRowCount(0)
        
        # Crear una instancia del filtro de eventos
        # self.ui.txtPlacaVehicular.textChanged.connect(self.fn_buscarDatosGuardados)
        self.ui.txtPlacaVehicular.textChanged.connect(self.fn_uppercase_filter)
        self.ui.txtCliente.textChanged.connect(self.fn_uppercase_filter)
        self.ui.txtConductor.textChanged.connect(self.fn_uppercase_filter)
        self.ui.txtProducto.textChanged.connect(self.fn_uppercase_filter)
        self.ui.txtTransportista.textChanged.connect(self.fn_uppercase_filter)
        self.ui.txtObservacion.textChanged.connect(self.fn_uppercase_filter)
        
        self.ui.txtPesoBruto.textChanged.connect(self.fn_entradaPesoBruto)
        self.ui.txtPesoTara.textChanged.connect(self.fn_entradaPesoBruto)
        
        self.ui.btn_salir.clicked.connect(self.fn_btnSalirDelSistema)
        self.ui.btn_minimizar.clicked.connect(self.fn_minimizarPrograma)
        self.ui.btnCancelar.clicked.connect(self.fn_btn_cancelar)
        self.ui.radioBtnIngreso.clicked.connect(self.fn_btn_ingreso)
        self.ui.radioBtnSalida.clicked.connect(self.fn_btn_salida)
        self.ui.btnCapturar.clicked.connect(self.fn_btn_capturar)
        self.ui.btnGuardar.clicked.connect(self.fn_btn_guardar)
        self.ui.btnImprimir.clicked.connect(self.fn_btn_imprimir)
        self.ui.btnActualizar.clicked.connect(self.fn_btn_actualizar)
        self.ui.btnCerrar.clicked.connect(self.fn_cerrarPrograma)
        self.ui.btnminimizar.clicked.connect(self.fn_minimizarPrograma)
        self.ui.btnNuevoTicket.clicked.connect(self.fn_btn_ingreso)
        #self.ui.btnBuscarTara.clicked.connect(self.fn_traer_tara)
        self.ui.btnCerrarFrmAlerta.clicked.connect(self.fn_btnCerrarFrmAlerta)
        self.ui.btnEditBruto.clicked.connect(self.fn_btnEditBruto)
        self.ui.btnEditTara.clicked.connect(self.fn_btnEditTara)
        self.ui.tblDetallePesadas.itemClicked.connect(self.fn_fila_seleccionada)
        self.ui.txtFiltraPlaca.textChanged.connect(self.fn_filtroPlaca)
        
        self.ui.frmAlerta.setHidden(True)
        self.ui.frmAlerta_2.setHidden(True)
        
        self.ui.btnCapturar.setEnabled(False)
        self.ui.btnCancelar.setEnabled(False)
        self.ui.btnImprimir.setEnabled(False)
        self.ui.btnGuardar.setEnabled(False)
        
        self.ui.btnReporte.clicked.connect(self.fn_abrirReporte)
        
    def fn_cerrarPrograma(self):
        self.close()

    def fn_btnSalirDelSistema(self):
        self.close()

    def fn_minimizarPrograma(self):
        self.showMinimized()
        
    def fn_btnCerrarFrmAlerta(self):
        self.ui.frmAlerta_2.setHidden(True)
        
    def fn_btnEditBruto(self):
        if(capturarPesoBruto or capturarPesoTara):
            self.ui.txtPesoBruto.setEnabled(True)
            self.ui.txtPesoBruto.setFocus(True)
        
    def fn_btnEditTara(self):
        if(capturarPesoBruto or capturarPesoTara):
            self.ui.txtPesoTara.setEnabled(True)
            self.ui.txtPesoTara.setFocus(True)
    
    def fn_filtroPlaca(self):
        filtro = self.ui.txtFiltraPlaca.text()

        for fila in range(self.ui.tblDetallePesadas.rowCount()):
            item = self.ui.tblDetallePesadas.item(fila, 1)

            if filtro.lower() in item.text().lower():
                self.ui.tblDetallePesadas.setRowHidden(fila, False)
            else:
                self.ui.tblDetallePesadas.setRowHidden(fila, True)
        
    def fn_uppercase_filter(self):
        sender = self.sender()
        if isinstance(sender, QtWidgets.QLineEdit):
            cursor_position = sender.cursorPosition()  # Guardar la posición del cursor
            text = sender.text()
            upper_text = text.upper()
            sender.setText(upper_text)
            # Restaurar la posición del cursor
            if cursor_position <= len(upper_text):
                sender.setCursorPosition(cursor_position)
            else:
                sender.setCursorPosition(len(upper_text))
                
    def fn_entradaPesoBruto(self):
        texto = self.ui.txtPesoBruto.text()
        #print("texto de la cadena",texto)
        cursor_position = self.ui.txtPesoBruto.cursorPosition()  # Guardar la posición del cursor

        nuevo_texto = ""
        punto_encontrado = False
    
        for char in texto:
            if char.isdigit():
                nuevo_texto += char
            elif char == '.' and not punto_encontrado:
                nuevo_texto += char
                punto_encontrado = True
    
        #print("texto valido", nuevo_texto)
    
        self.ui.txtPesoBruto.setText(nuevo_texto)
        self.ui.txtPesoBruto.setCursorPosition(cursor_position)
        
        self.fn_calcular_peso_neto()
    
    def fn_entradaPesoTara(self):
        texto = self.ui.txtPesoTara.text()
        cursor_position = self.ui.txtPesoTara.cursorPosition()  # Guardar la posición del cursor
        texto_valido = ''.join(filter(str.isdigit, texto))
        self.ui.txtPesoTara.setText(texto_valido)
        self.ui.txtPesoTara.setCursorPosition(cursor_position)
        
        self.fn_calcular_peso_neto()
        
    def fn_abrir_calendario(self):
        if(capturarPesoBruto or capturarPesoTara):
            if self.ui.calendarWidget.isVisible():
                self.ui.calendarWidget.setHidden(True)
            else:
                self.ui.calendarWidget.setHidden(False)
                selected_date = self.ui.dateEditConsultar.date()
                self.ui.calendarWidget.setSelectedDate(selected_date)
        
    def update_date(self, date):
        self.ui.dateEditConsultar.setDate(date)
        self.ui.calendarWidget.setHidden(True)
        self.ui.txtFiltraPlaca.setText("")
        fecha_qdate = self.ui.dateEditConsultar.date()
        fechaActual = fecha_qdate.toString('dd-MM-yyyy')
        self.fn_listarTabla(fechaActual)

    def evt_actualizar_peso(self, result):
        self.ui.lblDatosCablesSeriales.setText(result)  
        try:
            import re
            numeros_encontrados = re.findall(r"-?\d+(?:\.\d+)?", result) 

            if numeros_encontrados:  
                numero = numeros_encontrados[0]
                numero = numero.lstrip("0") if numero.lstrip("0") else "0"
                self.ui.lblPesoIndicador.setText(numero) 
            else:
                self.ui.lblPesoIndicador.setText("-----") 
        except Exception as e:
            print(f"Error al procesar el peso: {e}")
            self.ui.lblPesoIndicador.setText("-----")


    def evt_actualizar_estado(self, result):
        if (result == "0"):
            self.ui.lblEstadoIndicador.setText("FUERA DE LINEA")
            self.ui.lblEstadoIndicador.setStyleSheet("background-color: rgb(234, 29, 49);color: rgb(255, 255, 255);border-radius: 5px;")
        elif (result == "1"):
            self.ui.lblEstadoIndicador.setText("EN LINEA")
            self.ui.lblEstadoIndicador.setStyleSheet("background-color: rgb(32, 176, 20);color: rgb(255, 255, 255);border-radius: 5px;")
        
    def mostrar_fecha_hora(self,val):
        self.ui.lblFechaHora.setText(val)
    
    def fn_asigna_puertos_default(self):
        global COM1
        global COMAR

        puertos_default = self.conexion.db_consulta_puertos_default()
        
        COM1 = str(puertos_default[1])
        COMAR = str(puertos_default[2])
        
    def fn_asigna_nombreImpresora(self):
        global nombreImpresora
        global nombreEmpresa
        global ubicacionEmpresa
        global rucEmpresa

        datos_impresora = self.conexion.db_consulta_datosImpresora()
        
        nombreImpresora = str(datos_impresora[1])
        nombreEmpresa = str(datos_impresora[2])
        ubicacionEmpresa = str(datos_impresora[3])
        rucEmpresa = str(datos_impresora[4])
    
    def fn_asigna_recorte_cadena(self):
        global cadenaPrincipio
        global cadenaFinal

        cadena_recorte = self.conexion.db_consulta_recorte_cadena()
        
        cadenaPrincipio = int(cadena_recorte[1])
        cadenaFinal = int(cadena_recorte[2])
        
    def fn_limpiar_campos(self):
        self.ui.txtNroTicket.setText("")
        self.ui.txtFechaInicial.setText("")
        self.ui.txtHoraInicial.setText("")
        self.ui.txtFechaFinal.setText("")
        self.ui.txtHoraFinal.setText("")
        self.ui.txtPesoBruto.setText("")
        self.ui.txtPesoTara.setText("")
        self.ui.txtPesoNeto.setText("")        
        self.ui.txtPlacaVehicular.setText("")
        self.ui.txtCliente.setText("")
        self.ui.txtConductor.setText("")
        self.ui.txtProducto.setText("")
        self.ui.txtTransportista.setText("")
        self.ui.txtObservacion.setText("")
        
    def fn_btn_ingreso(self):
        global capturarPesoBruto
        
        capturarPesoBruto = True
        
        self.fn_limpiar_campos()
        self.fn_volver_colores()
        
        self.ui.txtPlacaVehicular.setEnabled(True)
        self.ui.txtCliente.setEnabled(True)
        self.ui.txtConductor.setEnabled(True)
        self.ui.txtProducto.setEnabled(True)
        self.ui.txtTransportista.setEnabled(True)
        self.ui.txtObservacion.setEnabled(True)
        self.ui.txtPesoBruto.setEnabled(False)
        self.ui.txtPesoTara.setEnabled(False)
        self.ui.txtFiltraPlaca.setEnabled(True)
        
        self.ui.btnCapturar.setEnabled(True)
        self.ui.btnCancelar.setEnabled(True)
        self.ui.btnImprimir.setEnabled(True)
        self.ui.btnGuardar.setEnabled(True)
        
        self.ui.txtPlacaVehicular.setFocus(True)
        self.ui.imgCancelar.setPixmap(QPixmap("Resources/cancelar.png"))
        self.ui.imgCamion.setPixmap(QPixmap("Resources/camion.png"))
        self.ui.imgGuardar.setPixmap(QPixmap("Resources/guardar.png"))
        self.ui.imgImprimir.setPixmap(QPixmap("Resources/impresora.png"))
        self.ui.lblGuardarActualizar.setText("Guardar")
        
        self.ui.btnCapturar.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.ui.btnCancelar.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.ui.btnGuardar.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.ui.btnImprimir.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        
        self.ui.radioBtnIngreso.setEnabled(False)
        self.ui.radioBtnSalida.setEnabled(False)
        
        ultimoTicket = self.conexion.db_consulta_ultimoTicket()
        ultimoTicket += 1
        formattedTicket = str(ultimoTicket).zfill(6)
        self.ui.txtNroTicket.setText(formattedTicket)
        
        self.ui.imgRefresh.setHidden(False)
        self.ui.btnNuevoTicket.setHidden(False)
        self.ui.lblNuevoTicket.setHidden(False)
        
        fechaActual = datetime.now().date()
        qdate = QDate(fechaActual.year, fechaActual.month, fechaActual.day)
        self.ui.dateEditConsultar.setDate(qdate)
        self.ui.txtFiltraPlaca.setText("")
        fechaActual = datetime.now().strftime('%d-%m-%Y')
        self.fn_listarTabla(fechaActual)
        
    def fn_btn_salida(self):
        global capturarPesoTara
        
        capturarPesoTara = True
        
        self.fn_limpiar_campos()
        
        self.ui.txtPlacaVehicular.setEnabled(True)
        self.ui.txtCliente.setEnabled(True)
        self.ui.txtConductor.setEnabled(True)
        self.ui.txtProducto.setEnabled(True)
        self.ui.txtTransportista.setEnabled(True)
        self.ui.txtObservacion.setEnabled(True)
        self.ui.txtPesoBruto.setEnabled(False)
        self.ui.txtPesoTara.setEnabled(False)
        self.ui.txtFiltraPlaca.setEnabled(True)
        
        self.ui.btnCapturar.setEnabled(True)
        self.ui.btnCancelar.setEnabled(True)
        self.ui.btnImprimir.setEnabled(True)
        self.ui.btnGuardar.setEnabled(True)
        
        self.ui.txtPlacaVehicular.setFocus(True)
        self.ui.imgCancelar.setPixmap(QPixmap("Resources/cancelar.png"))
        self.ui.imgCamion.setPixmap(QPixmap("Resources/camion.png"))
        self.ui.imgGuardar.setPixmap(QPixmap("Resources/guardar.png"))
        self.ui.imgImprimir.setPixmap(QPixmap("Resources/impresora.png"))
        self.ui.lblGuardarActualizar.setText("Actualizar")
        
        self.ui.btnCapturar.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.ui.btnCancelar.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.ui.btnGuardar.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.ui.btnImprimir.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        
        self.ui.radioBtnIngreso.setEnabled(False)
        self.ui.radioBtnSalida.setEnabled(False)
        
        fechaActual = datetime.now().date()
        qdate = QDate(fechaActual.year, fechaActual.month, fechaActual.day)
        self.ui.dateEditConsultar.setDate(qdate)
        self.ui.txtFiltraPlaca.setText("")
        fechaActual = datetime.now().strftime('%d-%m-%Y')
        self.fn_listarTabla(fechaActual)
        
    def fn_btn_cancelar(self):
        global capturarPesoBruto
        global capturarPesoTara
        
        capturarPesoBruto = False
        capturarPesoTara = False
        
        self.fn_limpiar_campos()
        self.fn_volver_colores()
        
        self.ui.txtPlacaVehicular.setEnabled(False)
        self.ui.txtCliente.setEnabled(False)
        self.ui.txtConductor.setEnabled(False)
        self.ui.txtProducto.setEnabled(False)
        self.ui.txtTransportista.setEnabled(False)
        self.ui.txtObservacion.setEnabled(False)
        self.ui.txtPesoBruto.setEnabled(False)
        self.ui.txtPesoTara.setEnabled(False)
        self.ui.txtFiltraPlaca.setEnabled(False)
        
        self.ui.btnCapturar.setEnabled(False)
        self.ui.btnCancelar.setEnabled(False)
        self.ui.btnImprimir.setEnabled(False)
        self.ui.btnGuardar.setEnabled(False)
        
        self.ui.txtPlacaVehicular.setFocus(True)
        self.ui.imgCancelar.setPixmap(QPixmap("Resources/cancelar_bn.png"))
        self.ui.imgCamion.setPixmap(QPixmap("Resources/camion_bn.png"))
        self.ui.imgGuardar.setPixmap(QPixmap("Resources/guardar_bn.png"))
        self.ui.imgImprimir.setPixmap(QPixmap("Resources/impresora_bn.png"))
        self.ui.lblGuardarActualizar.setText("Guardar")
        
        self.ui.btnCapturar.setCursor(QtGui.QCursor(QtCore.Qt.ArrowCursor))
        self.ui.btnCancelar.setCursor(QtGui.QCursor(QtCore.Qt.ArrowCursor))
        self.ui.btnGuardar.setCursor(QtGui.QCursor(QtCore.Qt.ArrowCursor))
        self.ui.btnImprimir.setCursor(QtGui.QCursor(QtCore.Qt.ArrowCursor))
        
        tablaDePesos = self.ui.tblDetallePesadas
        tablaDePesos.clearContents()
        tablaDePesos.setRowCount(0)

        self.ui.radioBtnIngreso.setCheckable(False)
        self.ui.radioBtnIngreso.setCheckable(True)
        self.ui.radioBtnIngreso.setChecked(False)
        
        self.ui.radioBtnSalida.setCheckable(False)
        self.ui.radioBtnSalida.setCheckable(True)
        self.ui.radioBtnSalida.setChecked(False)
        
        self.ui.radioBtnIngreso.setEnabled(True)
        self.ui.radioBtnSalida.setEnabled(True)
        
        self.ui.imgRefresh.setHidden(True)
        self.ui.btnNuevoTicket.setHidden(True)
        self.ui.lblNuevoTicket.setHidden(True)
        
        tablaDePesos = self.ui.tblDetallePesadas
        tablaDePesos.clearContents()
        tablaDePesos.setRowCount(0)
        
    def fn_btn_capturar(self):
        fechaActual = datetime.now().strftime('%d-%m-%Y')
        horaActual = datetime.now().strftime('%H:%M:%S')
        
        if(capturarPesoBruto):
            pesoBruto = self.ui.lblPesoIndicador.text().strip()
            if pesoBruto == "-----":
                pesoBruto = ""
            self.ui.txtPesoBruto.setText(pesoBruto)
            self.ui.txtFechaInicial.setText(fechaActual)
            self.ui.txtHoraInicial.setText(horaActual)
        elif (capturarPesoTara):
            pesoTara = self.ui.lblPesoIndicador.text().strip()
            if pesoTara == "-----":
                pesoTara = ""
            self.ui.txtPesoTara.setText(pesoTara)
            self.ui.txtFechaFinal.setText(fechaActual)
            self.ui.txtHoraFinal.setText(horaActual)
            
        self.ui.txtPesoBruto.setEnabled(False)
        self.ui.txtPesoTara.setEnabled(False)        
        self.fn_calcular_peso_neto()
            
    def fn_calcular_peso_neto(self):
        pesoBruto = self.ui.txtPesoBruto.text().strip()
        pesoTara = self.ui.txtPesoTara.text().strip()
        
        #print("peso bruto: ",pesoBruto,"peso Tara: ",pesoTara)

        try:
            pesoNeto = int(pesoBruto)-int(pesoTara)
            #print("PESO neto: ",pesoNeto)
            if float(pesoNeto) < 0:
                pesoNeto = int(pesoNeto)*-1
            self.ui.txtPesoNeto.setText(str(pesoNeto))
        except Exception as e:
            self.ui.txtPesoNeto.setText("")
        
    def fn_buscarDatosGuardados(self):
        cursor_position = self.ui.txtPlacaVehicular.cursorPosition()  # Guardar la posición del cursor
        text = self.ui.txtPlacaVehicular.text()
        upper_text = text.upper()
        self.ui.txtPlacaVehicular.setText(upper_text)
        self.ui.txtPlacaVehicular.setCursorPosition(len(upper_text))
        # Restaurar la posición del cursor
        if cursor_position <= len(upper_text):
            self.ui.txtPlacaVehicular.setCursorPosition(cursor_position)
        else:
            self.ui.txtPlacaVehicular.setCursorPosition(len(upper_text))
        
        valor = self.ui.txtPlacaVehicular.text()

        if (valor != "" and len(valor) >= 1):

            placaClienteSeleccionar = self.conexion.db_buscaDatosGuardados(valor)

            if placaClienteSeleccionar is not None:
                try:
                    placaClienteSeleccionar = placaClienteSeleccionar[0]
                    
                    self.ui.txtCliente.setText("" if placaClienteSeleccionar[1] is None else str(placaClienteSeleccionar[1]))
                    self.ui.txtConductor.setText("" if placaClienteSeleccionar[2] is None else str(placaClienteSeleccionar[2]))
                    self.ui.txtTransportista.setText("" if placaClienteSeleccionar[3] is None else str(placaClienteSeleccionar[3]))
                except Exception as e:
                    self.ui.txtCliente.setText("")
                    self.ui.txtConductor.setText("")
                    self.ui.txtTransportista.setText("")
            else:
                self.ui.txtCliente.setText("")
                self.ui.txtConductor.setText("")
                self.ui.txtTransportista.setText("")
        else:
            self.ui.txtCliente.setText("")
            self.ui.txtConductor.setText("")
            self.ui.txtTransportista.setText("")
            
    def fn_btn_guardar(self):
        try:
            txtNroTicket = str(int(self.ui.txtNroTicket.text().strip()))
        except Exception as e:
            pass
        txtFechaInicial = self.ui.txtFechaInicial.text().strip()
        txtHoraInicial = self.ui.txtHoraInicial.text().strip()
        txtFechaFinal = self.ui.txtFechaFinal.text().strip()
        txtHoraFinal = self.ui.txtHoraFinal.text().strip()
        txtPesoBruto = self.ui.txtPesoBruto.text().strip()
        txtPesoTara = self.ui.txtPesoTara.text().strip()      
        txtPlacaVehicular = self.ui.txtPlacaVehicular.text().strip()
        txtCliente = self.ui.txtCliente.text().strip()
        txtConductor = self.ui.txtConductor.text().strip()
        txtProducto = self.ui.txtProducto.text().strip()
        txtTransportista = self.ui.txtTransportista.text().strip()
        txtObservacion = self.ui.txtObservacion.text().strip()
        
        self.ui.txtPesoBruto.setEnabled(False)
        self.ui.txtPesoTara.setEnabled(False)
        
        if txtPlacaVehicular.strip() == '' or txtCliente.strip() == '' or txtProducto.strip() == '' or txtConductor.strip() == '' or txtPesoBruto.strip() == '':
            self.fn_alerta("ADVENTENCIA!", error, "Por favor, complete los campos obligatorios.")
            
            if self.ui.txtPesoBruto.text().strip() == '':
                self.ui.txtPesoBruto.setStyleSheet("background-color: rgb(255, 244, 209);color: rgb(0, 0, 0);border: 2px solid #FF0000;border-radius: 5px;padding-left: 10px;padding-right: 10px;")
            else:
                self.ui.txtPesoBruto.setStyleSheet("background-color: rgb(255, 244, 209);color: rgb(0, 0, 0);border: 2px solid #FFCC21;border-radius: 5px;padding-left: 10px;padding-right: 10px;")
            
            if self.ui.txtPlacaVehicular.text().strip() == '':
                self.ui.txtPlacaVehicular.setStyleSheet("background-color: rgb(255, 244, 209);color: rgb(0, 0, 0);border: 2px solid #FF0000;border-radius: 5px;padding-left: 10px;padding-right: 10px;")
            else:
                self.ui.txtPlacaVehicular.setStyleSheet("background-color: rgb(255, 244, 209);color: rgb(0, 0, 0);border: 2px solid #FFCC21;border-radius: 5px;padding-left: 10px;padding-right: 10px;")

            if self.ui.txtCliente.text().strip() == '':
                self.ui.txtCliente.setStyleSheet("background-color: rgb(255, 244, 209);color: rgb(0, 0, 0);border: 2px solid #FF0000;border-radius: 5px;padding-left: 10px;padding-right: 10px;")
            else:
                self.ui.txtCliente.setStyleSheet("background-color: rgb(255, 244, 209);color: rgb(0, 0, 0);border: 2px solid #FFCC21;border-radius: 5px;padding-left: 10px;padding-right: 10px;")

            if self.ui.txtConductor.text().strip() == '':
                self.ui.txtConductor.setStyleSheet("background-color: rgb(255, 244, 209);color: rgb(0, 0, 0);border: 2px solid #FF0000;border-radius: 5px;padding-left: 10px;padding-right: 10px;")
            else:
                self.ui.txtConductor.setStyleSheet("background-color: rgb(255, 244, 209);color: rgb(0, 0, 0);border: 2px solid #FFCC21;border-radius: 5px;padding-left: 10px;padding-right: 10px;")

            if self.ui.txtProducto.text().strip() == '':
                self.ui.txtProducto.setStyleSheet("background-color: rgb(255, 244, 209);color: rgb(0, 0, 0);border: 2px solid #FF0000;border-radius: 5px;padding-left: 10px;padding-right: 10px;")
            else:
                self.ui.txtProducto.setStyleSheet("background-color: rgb(255, 244, 209);color: rgb(0, 0, 0);border: 2px solid #FFCC21;border-radius: 5px;padding-left: 10px;padding-right: 10px;")
        else:
            if(capturarPesoBruto):
                respuesta = self.conexion.db_registrarPesadas(txtNroTicket,txtFechaInicial,txtHoraInicial,txtFechaFinal,txtHoraFinal,txtPesoBruto,txtPesoTara,txtPlacaVehicular,txtCliente,txtConductor,txtProducto,txtTransportista,txtObservacion)

                if(respuesta):
                    self.fn_alerta("¡REGISTRO EXITOSO!",correcto,"El registro se realizo correctamente.")
                    self.ui.txtFiltraPlaca.setText("")
                    fechaActual = datetime.now().strftime('%d-%m-%Y')
                    self.fn_listarTabla(fechaActual)
                else:
                    self.ui.lblAlertaTitulo_2.setStyleSheet("color: #EA1D31")
                    self.ui.lblAlertaTexto_2.setStyleSheet("font-size:16pt;")
                    self.ui.frmAlerta_2.setHidden(False)
                    self.ui.lblAlertaTitulo_2.setText("¡ADVERTENCIA!")
                    self.ui.imgIconAlerta_2.setPixmap(QPixmap(error))
                    self.ui.lblAlertaTexto_2.setText("Este ticket ya ha sido registrado. \n ¿Desea actualizar el ticket?")
                    
                self.fn_actualizar_ingresar_datos()
            elif (capturarPesoTara):
                self.fn_btn_actualizar()
        
    def fn_alerta(self,titulo,imagen,mensaje,tiempo = 1000):
        if imagen == correcto:
            self.ui.lblAlertaTitulo.setStyleSheet("color: #24D315")
            self.ui.lblAlertaTexto.setStyleSheet("font-size:16pt;")
        elif imagen == error:
            self.ui.lblAlertaTitulo.setStyleSheet("color: #EA1D31")
            self.ui.lblAlertaTexto.setStyleSheet("font-size:16pt;")
        self.ui.frmAlerta.setHidden(False)
        self.ui.lblAlertaTitulo.setText(titulo)
        self.ui.imgIconAlerta.setPixmap(QPixmap(imagen))
        self.ui.lblAlertaTexto.setText(mensaje)

        timer = QtCore.QTimer()
        timer.singleShot(tiempo, lambda: self.ui.frmAlerta.setHidden(True))
        
    def fn_volver_colores(self):
        self.ui.txtProducto.setStyleSheet("background-color: rgb(255, 244, 209);color: rgb(0, 0, 0);border: 2px solid #FFCC21;border-radius: 5px;padding-left: 10px;padding-right: 10px;")
        self.ui.txtPlacaVehicular.setStyleSheet("background-color: rgb(255, 244, 209);color: rgb(0, 0, 0);border: 2px solid #FFCC21;border-radius: 5px;padding-left: 10px;padding-right: 10px;")
        self.ui.txtCliente.setStyleSheet("background-color: rgb(255, 244, 209);color: rgb(0, 0, 0);border: 2px solid #FFCC21;border-radius: 5px;padding-left: 10px;padding-right: 10px;")
        self.ui.txtConductor.setStyleSheet("background-color: rgb(255, 244, 209);color: rgb(0, 0, 0);border: 2px solid #FFCC21;border-radius: 5px;padding-left: 10px;padding-right: 10px;")
        self.ui.txtPesoBruto.setStyleSheet("background-color: rgb(255, 244, 209);color: rgb(0, 0, 0);border: 2px solid #FFCC21;border-radius: 5px;padding-left: 10px;padding-right: 10px;")
    
    def fn_traer_tara(self):
        fechaActual = datetime.now().strftime('%d-%m-%Y')
        horaActual = datetime.now().strftime('%H:%M:%S')
        
        if(capturarPesoBruto or capturarPesoTara):
            if self.ui.txtPlacaVehicular.text().strip() == '':
                self.ui.txtPlacaVehicular.setStyleSheet("background-color: rgb(255, 244, 209);color: rgb(0, 0, 0);border: 2px solid #FF0000;border-radius: 5px;padding-left: 10px;padding-right: 10px;")
                self.ui.txtPlacaVehicular.setFocus(True)
            else:
                self.ui.txtPlacaVehicular.setStyleSheet("background-color: rgb(255, 244, 209);color: rgb(0, 0, 0);border: 2px solid #FFCC21;border-radius: 5px;padding-left: 10px;padding-right: 10px;")
                
                valor = self.ui.txtPlacaVehicular.text().strip()
                placaClienteSeleccionar = self.conexion.db_busca_tara_guardada(valor)

                if placaClienteSeleccionar is not None:
                    try:
                        placaClienteSeleccionar = placaClienteSeleccionar[0]

                        self.ui.txtPesoTara.setText("" if placaClienteSeleccionar[0] is None else str(placaClienteSeleccionar[0]))
                        txtPesoTara = self.ui.txtPesoTara.text()
                        if txtPesoTara != "":
                            self.ui.txtFechaFinal.setText(fechaActual)
                            self.ui.txtHoraFinal.setText(horaActual)
                    except Exception as e:
                        self.ui.txtPesoTara.setText("")
                        
                    self.fn_calcular_peso_neto()
    
    def fn_btn_actualizar(self):
        self.ui.frmAlerta_2.setHidden(True)
        try:
            txtNroTicket = str(int(self.ui.txtNroTicket.text().strip()))
        except Exception as e:
            pass
        
        if txtNroTicket != "":
            txtFechaInicial = self.ui.txtFechaInicial.text().strip()
            txtHoraInicial = self.ui.txtHoraInicial.text().strip()
            txtFechaFinal = self.ui.txtFechaFinal.text().strip()
            txtHoraFinal = self.ui.txtHoraFinal.text().strip()
            txtPesoBruto = self.ui.txtPesoBruto.text().strip()
            txtPesoTara = self.ui.txtPesoTara.text().strip()      
            txtPlacaVehicular = self.ui.txtPlacaVehicular.text().strip()
            txtCliente = self.ui.txtCliente.text().strip()
            txtConductor = self.ui.txtConductor.text().strip()
            txtProducto = self.ui.txtProducto.text().strip()
            txtTransportista = self.ui.txtTransportista.text().strip()
            txtObservacion = self.ui.txtObservacion.text().strip()
            
            self.ui.txtPesoBruto.setEnabled(False)
            self.ui.txtPesoTara.setEnabled(False)
            
            if txtPlacaVehicular.strip() == '' or txtCliente.strip() == '' or txtProducto.strip() == '' or txtConductor.strip() == '' or txtPesoBruto.strip() == '':
                self.fn_alerta("ADVENTENCIA!", error, "Por favor, complete los campos obligatorios.")
                
                if self.ui.txtPesoBruto.text().strip() == '':
                    self.ui.txtPesoBruto.setStyleSheet("background-color: rgb(255, 244, 209);color: rgb(0, 0, 0);border: 2px solid #FF0000;border-radius: 5px;padding-left: 10px;padding-right: 10px;")
                else:
                    self.ui.txtPesoBruto.setStyleSheet("background-color: rgb(255, 244, 209);color: rgb(0, 0, 0);border: 2px solid #FFCC21;border-radius: 5px;padding-left: 10px;padding-right: 10px;")
                
                if self.ui.txtPlacaVehicular.text().strip() == '':
                    self.ui.txtPlacaVehicular.setStyleSheet("background-color: rgb(255, 244, 209);color: rgb(0, 0, 0);border: 2px solid #FF0000;border-radius: 5px;padding-left: 10px;padding-right: 10px;")
                else:
                    self.ui.txtPlacaVehicular.setStyleSheet("background-color: rgb(255, 244, 209);color: rgb(0, 0, 0);border: 2px solid #FFCC21;border-radius: 5px;padding-left: 10px;padding-right: 10px;")

                if self.ui.txtCliente.text().strip() == '':
                    self.ui.txtCliente.setStyleSheet("background-color: rgb(255, 244, 209);color: rgb(0, 0, 0);border: 2px solid #FF0000;border-radius: 5px;padding-left: 10px;padding-right: 10px;")
                else:
                    self.ui.txtCliente.setStyleSheet("background-color: rgb(255, 244, 209);color: rgb(0, 0, 0);border: 2px solid #FFCC21;border-radius: 5px;padding-left: 10px;padding-right: 10px;")

                if self.ui.txtConductor.text().strip() == '':
                    self.ui.txtConductor.setStyleSheet("background-color: rgb(255, 244, 209);color: rgb(0, 0, 0);border: 2px solid #FF0000;border-radius: 5px;padding-left: 10px;padding-right: 10px;")
                else:
                    self.ui.txtConductor.setStyleSheet("background-color: rgb(255, 244, 209);color: rgb(0, 0, 0);border: 2px solid #FFCC21;border-radius: 5px;padding-left: 10px;padding-right: 10px;")

                if self.ui.txtProducto.text().strip() == '':
                    self.ui.txtProducto.setStyleSheet("background-color: rgb(255, 244, 209);color: rgb(0, 0, 0);border: 2px solid #FF0000;border-radius: 5px;padding-left: 10px;padding-right: 10px;")
                else:
                    self.ui.txtProducto.setStyleSheet("background-color: rgb(255, 244, 209);color: rgb(0, 0, 0);border: 2px solid #FFCC21;border-radius: 5px;padding-left: 10px;padding-right: 10px;")
            else:
                respuesta = self.conexion.db_actualizarPesadas(txtNroTicket,txtFechaInicial,txtHoraInicial,txtFechaFinal,txtHoraFinal,txtPesoBruto,txtPesoTara,txtPlacaVehicular,txtCliente,txtConductor,txtProducto,txtTransportista,txtObservacion)

                if(respuesta):
                    self.fn_alerta("¡ACTUALIZACIÓN EXITOSA!",correcto,"El registro se actualizo correctamente.")
                    self.ui.txtFiltraPlaca.setText("")
                    fecha_qdate = self.ui.dateEditConsultar.date()
                    fechaActual = fecha_qdate.toString('dd-MM-yyyy')
                    self.fn_listarTabla(fechaActual)
                else:
                    self.fn_alerta("¡ERROR!",error,"Ocurrio un ERROR al actualizar los datos.",2000)
                    
                self.fn_actualizar_ingresar_datos()
                
    def fn_actualizar_ingresar_datos(self):
        txtPesoTara = self.ui.txtPesoTara.text().strip()      
        txtPlacaVehicular = self.ui.txtPlacaVehicular.text().strip()
        txtCliente = self.ui.txtCliente.text().strip()
        txtConductor = self.ui.txtConductor.text().strip()
        txtTransportista = self.ui.txtTransportista.text().strip()
        
        resultado = self.conexion.db_verificarPlacaExistente(txtPlacaVehicular)
        
        if resultado is not None and len(resultado) > 0:             
            self.conexion.db_actualizarDatosGuardados(txtPlacaVehicular,txtCliente,txtConductor,txtTransportista)
            if txtPesoTara != "":
                self.conexion.db_actualizarDatosGuardadosTara(txtPesoTara)
        else:
            self.conexion.db_registrarDatosGuardados(txtPlacaVehicular,txtCliente,txtConductor,txtTransportista,txtPesoTara)
            
    def fn_listarTabla(self, fecha):
        tablaDePesos = self.ui.tblDetallePesadas
        tablaDePesos.clearContents()
        tablaDePesos.setRowCount(0)
        
        pesosListarTabla = self.conexion.db_listarPesosTabla(fecha)
        
        if pesosListarTabla != "" and pesosListarTabla != None:
        
            if len(pesosListarTabla) > 0:
            
                for row_number, row_data in enumerate(pesosListarTabla):
                    
                        tablaDePesos.insertRow(row_number)
                        
                        for column_number, data in enumerate(row_data):
                            if column_number == 0:
                                data = str(data).zfill(5)
                            
                            item = QTableWidgetItem(str(data))
                            item.setTextAlignment(Qt.AlignCenter)
                            tablaDePesos.setItem(row_number, column_number, item)
                        
                        if (row_data[5] != ""):
                            self.fn_pintarCeldasConTara(row_number)
                            
    def fn_pintarCeldasConTara(self, row):
        tablaDePesos = self.ui.tblDetallePesadas
        for e in range(tablaDePesos.columnCount()):
            item = tablaDePesos.item(row, e)
            item.setBackground(QColor(0, 170, 0))
            item.setForeground(QColor(255, 255, 255))
                            
    def fn_fila_seleccionada(self, item):
        if(not capturarPesoBruto):
            # Obtener la fila seleccionada
            fila = item.row()
            dato = self.ui.tblDetallePesadas.item(fila, 0).text()
            self.fn_traerDatosAnteriores(dato)
        
    def fn_traerDatosAnteriores(self, dato):
        respuesta = self.conexion.db_traerDatosAnteriores(dato)
        try:
            respuesta = respuesta[0]
            nroTicket = str(respuesta[0]).zfill(6)
            self.ui.txtNroTicket.setText(str(nroTicket))
            self.ui.txtFechaInicial.setText(str(respuesta[1]))
            self.ui.txtHoraInicial.setText(str(respuesta[3]))
            self.ui.txtFechaFinal.setText(str(respuesta[2]))
            self.ui.txtHoraFinal.setText(str(respuesta[4]))
            self.ui.txtPesoBruto.setText(str(respuesta[5]))
            self.ui.txtPesoTara.setText(str(respuesta[6]))       
            self.ui.txtPlacaVehicular.setText(str(respuesta[7]))
            self.ui.txtCliente.setText(str(respuesta[8]))
            self.ui.txtConductor.setText(str(respuesta[9]))
            self.ui.txtProducto.setText(str(respuesta[10]))
            self.ui.txtTransportista.setText(str(respuesta[11]))
            self.ui.txtObservacion.setText(str(respuesta[12]))
            
            self.fn_calcular_peso_neto()
        except Exception as e:
            pass

    def fn_btn_imprimir(self):
        fecha_inicial_completa = self.ui.txtFechaInicial.text().strip()
        fecha_inicial_solo_fecha = fecha_inicial_completa.split(" ")[0] if fecha_inicial_completa else ""
    
        datos = {
            "txtNroTicket": self.ui.txtNroTicket.text().strip(),
            "txtFechaInicial": fecha_inicial_solo_fecha,
            "txtHoraInicial": self.ui.txtHoraInicial.text().strip(),
            "txtFechaFinal": self.ui.txtFechaFinal.text().strip(),
            "txtHoraFinal": self.ui.txtHoraFinal.text().strip(),
            "txtPesoBruto": self.ui.txtPesoBruto.text().strip(),
            "txtPesoTara": self.ui.txtPesoTara.text().strip(),
            "txtPesoNeto": self.ui.txtPesoNeto.text().strip(),
            "txtPlacaVehicular": self.ui.txtPlacaVehicular.text().strip(),
            "txtCliente": self.ui.txtCliente.text().strip(),
            "txtConductor": self.ui.txtConductor.text().strip(),
            "txtProducto": self.ui.txtProducto.text().strip(),
            "txtTransportista": self.ui.txtTransportista.text().strip(),
            "txtObservacion": self.ui.txtObservacion.text().strip()
        }

        self.crear_pdf(datos)
        

    def crear_pdf(self, datos, filename="output.pdf"):
        c = canvas.Canvas(filename, pagesize=letter)
        width, height = letter
        
        ejeY = 0

        ejeY += 40
        c.setFont("Helvetica-Bold", 14)
        c.drawString(50, height - ejeY, f"{nombreEmpresa}")
        ejeY += 20
        c.setFont("Helvetica", 12)
        c.drawString(50, height - ejeY, f"{ubicacionEmpresa}")
        ejeY += 20
        c.drawString(50, height - ejeY, f"R.U.C. : {rucEmpresa}")
        
        ejeY += 30
        c.setFont("Helvetica-Bold", 12)
        c.drawString(450, height - ejeY, f"Nro Ticket : {datos['txtNroTicket']}")
        ejeY += 30
        c.setFont("Helvetica", 12)
        c.drawString(50, height - ejeY, f"Placa Vehicular  : {datos['txtPlacaVehicular']}")
        # c.drawString(300, height - ejeY, f"Carreta : {datos['txtCarreta']}")  
        ejeY += 20 
        c.drawString(50, height - ejeY, f"Conductor          : {datos['txtConductor']}")
        ejeY += 20 
        c.drawString(50, height - ejeY, f"Transportista     : {datos['txtTransportista']}")
        ejeY += 20 
        c.drawString(50, height - ejeY, f"Razon Social     : {datos['txtCliente']}")
        ejeY += 20 
        c.drawString(50, height - ejeY, f"Producto            : {datos['txtProducto']}")
        ejeY += 20 
        # c.drawString(50, height - ejeY, f"Precio                : {datos['txtPrecio']}")
        # ejeY += 20 
        c.drawString(50, height - ejeY, f"Observacion      : {datos['txtObservacion']}")
        c.line(50, height - ejeY - 10, width - 50, height - ejeY - 10)
        ejeY += 30 
        c.drawString(50, height - ejeY, f"Fecha Inicial      : {datos['txtFechaInicial']}")
        c.drawString(300, height - ejeY, f"Hora Inicial : {datos['txtHoraInicial']}")
        ejeY += 20 
        c.drawString(50, height - ejeY, f"Fecha Final       : {datos['txtFechaFinal']}")
        c.drawString(300, height - ejeY, f"Hora Final  : {datos['txtHoraFinal']}")
        ejeY += 30 
        c.setFont("Helvetica-Bold", 14)
        c.drawString(50, height - ejeY, f"PESO BRUTO : {datos['txtPesoBruto']}")
        ejeY += 20 
        c.drawString(50, height - ejeY, f"PESO TARA   : {datos['txtPesoTara']}")
        ejeY += 20 
        # Dibujar el cuadrado y el texto "PESO NETO"
        c.drawString(55, height - ejeY, f"PESO NETO  : {datos['txtPesoNeto']}")
        neto_text_width = c.stringWidth(f"PESO NETO  : {datos['txtPesoNeto']}", "Helvetica-Bold", 14)
        rect_x = 50
        rect_y = height - ejeY - 5
        rect_width = neto_text_width + 10
        rect_height = 20
        c.rect(rect_x, rect_y, rect_width, rect_height, stroke=1, fill=0)
        
        # try:
        #     imagen_path = "resources/logoEmpresa.png"
        #     imagen_x = 410
        #     imagen_y = height - 400
        #     imagen_ancho = 150
        #     imagen_alto = 75

        #     # Dibujar la imagen en el lienzo
        #     c.drawImage(imagen_path, imagen_x, imagen_y, width=imagen_ancho, height=imagen_alto)
        # except Exception as e:
        #     pass

        c.save()

        self.previsualizar_pdf(filename)
        
    def previsualizar_pdf(self, filename):
        self.viewer = PDFViewer(filename)
        self.viewer.show()
        
    def fn_abrirReporte(self):
        dialogo_modal_reporte = QDialog(self)
        dialogo_modal_reporte.setWindowIcon(QtGui.QIcon("Resources/icono.png"))
        dialogo_modal_reporte.setWindowTitle('Sistema de Pesaje || Balinsa')
        self.modal_reporte.setupUi(dialogo_modal_reporte)
        self.modal_reporte.btn_exportar.clicked.connect(self.exportar_a_excel)

        fechaActual = datetime.now().date()
        qdate = QDate(fechaActual.year, fechaActual.month, fechaActual.day)
        self.modal_reporte.dateEditConsultarDesde.setDate(qdate)
        self.modal_reporte.dateEditConsultarHasta.setDate(qdate)
        self.modal_reporte.calendarWidgetDesde.setHidden(True)
        self.modal_reporte.calendarWidgetHasta.setHidden(True)

        self.modal_reporte.btnCalendarioDesde.clicked.connect(self.fn_abrir_calendario_reporteDesde)
        self.modal_reporte.calendarWidgetDesde.clicked.connect(self.update_date_reporteDesde)
        self.modal_reporte.imgCalendarioDesde.setPixmap(QPixmap("Resources/calendario.png"))
        self.modal_reporte.btnCalendarioHasta.clicked.connect(self.fn_abrir_calendario_reporteHasta)
        self.modal_reporte.calendarWidgetHasta.clicked.connect(self.update_date_reporteHasta)
        self.modal_reporte.imgCalendarioHasta.setPixmap(QPixmap("Resources/calendario.png"))

        self.modal_reporte.dateEditConsultarDesde.dateChanged.connect(self.actualizar_tabla_reporte)
        self.modal_reporte.dateEditConsultarHasta.dateChanged.connect(self.actualizar_tabla_reporte)

        self.actualizar_tabla_reporte()
        dialogo_modal_reporte.exec_()

    def fn_abrir_calendario_reporteDesde(self):
        if self.modal_reporte.calendarWidgetDesde.isVisible():
            self.modal_reporte.calendarWidgetDesde.setHidden(True)
        else:
            self.modal_reporte.calendarWidgetDesde.setHidden(False)
            selected_date = self.modal_reporte.dateEditConsultarDesde.date()
            self.modal_reporte.calendarWidgetDesde.setSelectedDate(selected_date)

    def update_date_reporteDesde(self, date):
        self.modal_reporte.dateEditConsultarDesde.setDate(date)
        self.modal_reporte.calendarWidgetDesde.setHidden(True)

    def fn_abrir_calendario_reporteHasta(self):
        if self.modal_reporte.calendarWidgetHasta.isVisible():
            self.modal_reporte.calendarWidgetHasta.setHidden(True)
        else:
            self.modal_reporte.calendarWidgetHasta.setHidden(False)
            selected_date = self.modal_reporte.dateEditConsultarHasta.date()
            self.modal_reporte.calendarWidgetHasta.setSelectedDate(selected_date)

    def update_date_reporteHasta(self, date):
        self.modal_reporte.dateEditConsultarHasta.setDate(date)
        self.modal_reporte.calendarWidgetHasta.setHidden(True)

    def actualizar_tabla_reporte(self):
        fecha_desde = datetime.strptime(self.modal_reporte.dateEditConsultarDesde.date().toString("dd-MM-yyyy"), "%d-%m-%Y")
        fecha_hasta = datetime.strptime(self.modal_reporte.dateEditConsultarHasta.date().toString("dd-MM-yyyy"), "%d-%m-%Y")

        datos = self.obtener_datos_pesadas(fecha_desde, fecha_hasta)

        if datos is None:
            print("Error al obtener datos de la base de datos. Por favor, revise los logs.")
            return

        if not datos:
            print("No se encontraron registros para las fechas seleccionadas.")
            self.modal_reporte.tblDetallePesadas.setRowCount(0)  # Limpiar tabla
            return
        
        self.modal_reporte.tblDetallePesadas.setRowCount(0)

        # Agregar datos
        for fila_datos in datos:
            row_position = self.modal_reporte.tblDetallePesadas.rowCount()
            self.modal_reporte.tblDetallePesadas.insertRow(row_position)
            for column, dato in enumerate(fila_datos):
                self.modal_reporte.tblDetallePesadas.setItem(row_position, column, QTableWidgetItem(str(dato)))


    def obtener_datos_pesadas(self, fecha_desde, fecha_hasta):
        try:
            if not self.conexion.conexionsql:
                print("La conexión a la base de datos no está disponible.")
                return None

            # Convertir las fechas
            fecha_desde_texto = fecha_desde.strftime("%d-%m-%Y")
            fecha_hasta_texto = fecha_hasta.strftime("%d-%m-%Y")

            query = """
                SELECT 
                    idPesada, fechaPesajeInicial, horaPesajeInicial, 
                    pesoBruto, pesoTara, 
                    placaVehicular, cliente, conductor, transportista, 
                    producto, observacion, 
                    fechaPesajeFinal, horaPesajeFinal
                FROM tbl_pesadas 
                WHERE STR_TO_DATE(fechaPesajeInicial, '%d-%m-%Y') 
                BETWEEN STR_TO_DATE(%s, '%d-%m-%Y') 
                AND STR_TO_DATE(%s, '%d-%m-%Y');
            """

            with self.conexion.conexionsql.cursor() as cursor:
                cursor.execute(query, (fecha_desde_texto, fecha_hasta_texto))
                resultados = cursor.fetchall()

            datos = []
            for fila in resultados:
                try:
                    peso_bruto = float(fila[3]) if fila[3] else 0
                    peso_tara = float(fila[4]) if fila[4] else 0
                    peso_neto = peso_bruto - peso_tara
                except ValueError:
                    peso_neto = 0 
                
                fecha_pesaje_inicial = str(fila[1]).split(" ")[0] if fila[1] else ""

                datos.append([
                    str(fila[0]),  # idPesada
                    fecha_pesaje_inicial,  # fechaPesajeInicial
                    fila[2] or "",  # horaPesajeInicial
                    str(peso_bruto),  # pesoBruto
                    str(peso_tara),  # pesoTara
                    str(peso_neto),  # pesoNeto (calculado en Python)
                    fila[5] or "",  # placaVehicular
                    fila[6] or "",  # cliente
                    fila[7] or "",  # conductor
                    fila[8] or "",  # transportista
                    fila[9] or "",  # producto
                    fila[10] or "",  # observacion
                    fila[11] or "",  # fechaPesajeFinal
                    fila[12] or "",  # horaPesajeFinal
                ])

            return datos

        except Exception as e:
            print(f"Error al obtener datos de pesadas: {e}")
            return None
        
    def exportar_a_excel(self):
        try:
            print("Iniciando exportación a Excel.")
            
            nombre_archivo, _ = QFileDialog.getSaveFileName(
                self, 
                "Guardar Reporte", 
                "Reporte_Pesadas.xlsx", 
                "Archivos Excel (*.xlsx)"
            )
            
            if not nombre_archivo: 
                print("Exportación cancelada por el usuario.")
                return

            workbook = openpyxl.Workbook()
            hoja = workbook.active
            hoja.title = "Reporte de Pesadas"
            print("Libro de Excel creado.")
            
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            thin_border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )
            
            headers = [
                "ID Pesada", "Fecha Pesaje Inicial", "Hora Pesaje Inicial", 
                "Peso Bruto", "Peso Tara", "Peso Neto", 
                "Placa Vehicular", "Cliente", "Conductor", 
                "Transportista", "Producto", "Observación", 
                "Fecha Pesaje Final", "Hora Pesaje Final"
            ]
            
            for col, header in enumerate(headers, start=1):
                cell = hoja.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
            print("Encabezados escritos con estilo.")

            row_count = self.modal_reporte.tblDetallePesadas.rowCount()
            column_count = self.modal_reporte.tblDetallePesadas.columnCount()
            print(f"Filas: {row_count}, Columnas: {column_count}")

            for row in range(row_count):
                for col in range(column_count):
                    valor = self.modal_reporte.tblDetallePesadas.item(row, col).text() if self.modal_reporte.tblDetallePesadas.item(row, col) else ""
                    cell = hoja.cell(row=row+2, column=col+1, value=valor)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal="left", vertical="center")
            
            for col in hoja.columns:
                max_length = 0
                col_letter = col[0].column_letter 
                for cell in col:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                adjusted_width = max_length + 2 
                hoja.column_dimensions[col_letter].width = adjusted_width
            print("Ancho de columnas ajustado.")

            workbook.save(nombre_archivo)
            print(f"Archivo Excel guardado como {nombre_archivo}.")

        except Exception as e:
            print(f"Error al exportar a Excel: {e}")
            QMessageBox.critical(self, "Error", f"No se pudo exportar el archivo. Detalles: {e}")

